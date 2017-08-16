# comment!
from openpyxl import load_workbook
import numpy as np
from acc_deep_flow import acc_deep_flow
from heel_strike import heel_strike
from PIL import Image
import os
from checkDir import checkDir
import matplotlib.pyplot as plt



def reNum(num, addZero=False):
	if addZero == True:
		if num < 10:
			num = '00'+str(num)
		elif num < 100:
			num = '0'+str(num)
		else:
			num = str(num)

	else:
		num = str(num)

	return num


def maskCoords(mask):
	maskShape = mask.shape

	if len(maskShape) == 2:
		coords = np.argwhere(mask == 255)
		# print coords
	else:
		coords = np.argwhere(mask == [255,255,255])
		# print coords

	return coords


def main(radAcc_thres, frame_path, rmBG_path, mask_path, save_path, given_obj_index = None):
	if given_obj_index != None:
		mask_path = os.path.join(mask_path, given_obj_index)

	for indata_path, indata_dirnames, data in os.walk(mask_path):
		data_folders = indata_path.split('/')
		obj_index = data_folders[-1]
		masks = os.listdir(indata_path)
		masks = filter(lambda x: x.lower().endswith('.png'), masks)


		for img in masks:
			frame_num = os.path.splitext(img)[0]
			print frame_num

			img1_path = os.path.join(rmBG_path, obj_index, img)
			img1 = np.array(Image.open(img1_path))
			[row, col, dim] = img1.shape
			
			img2_num = int(frame_num)+1
			img2_path = os.path.join(rmBG_path, obj_index, reNum(img2_num))+".png"

			while os.path.isfile(img2_path) == False:
				img2_num += 1
				img2_path = os.path.join(rmBG_path, obj_index, reNum(img2_num))+".png"
			img2 = np.array(Image.open(img2_path))


			img3_num = int(img2_num)+1
			img3_path = os.path.join(rmBG_path, obj_index, reNum(img3_num))+".png"

			while os.path.isfile(img3_path) == False:
				img3_num += 1
				img3_path = os.path.join(rmBG_path, obj_index, reNum(img3_num))+".png"

			img3 = np.array(Image.open(img3_path))

							
			mask1_path = os.path.join(indata_path, img)
			mask1 = np.array(Image.open(mask1_path))
			coords = maskCoords(mask1)
			# print coords.shape
				
			sil1R_min = min(coords[:,0])
			sil1R_max = max(coords[:,0])
			# print "sil1", sil1R_min, sil1R_max
				
			sil1C_min = min(coords[:,1])
			sil1C_max = max(coords[:,1])
			# print "sil1", sil1C_min, sil1C_max


			mask2_path = os.path.join(indata_path, reNum(img2_num))+".png"
			mask2 = np.array(Image.open(mask2_path))
			coords = maskCoords(mask2)
			# print coords

			sil2R_min = min(coords[:,0])
			sil2R_max = max(coords[:,0])
			# print "sil2", sil2R_min, sil2R_max

			sil2C_min= min(coords[:,1])
			sil2C_max = max(coords[:,1])
			# print "sil2", sil2C_min, sil2C_max
				
			mask3_path = os.path.join(indata_path, reNum(img3_num))+".png"
			mask3 = np.array(Image.open(mask3_path))
			coords = maskCoords(mask3)
			# print coords

			sil3R_min = min(coords[:,0])
			sil3R_max = max(coords[:,0])
			# print "sil3", sil3R_min, sil3R_max

			sil3C_min = min(coords[:,1])
			sil3C_max = max(coords[:,1])
			# print "sil3", sil3C_min, sil3C_max


			bound_L = min(sil1C_min, sil2C_min, sil3C_min) - 50
			if bound_L < 0:
				bound_L = 0
				
			bound_R = max(sil1C_max, sil2C_max, sil3C_max) + 50
			if bound_R >= col:
				bound_R = col-1


			bound_U = min(sil1R_min, sil2R_min, sil3R_min) - 50
			if bound_U < 0:
				bound_U = 0

			bound_D = max(sil1R_max, sil2R_max, sil3R_max) + 50
			if bound_D >= row:
				bound_D = row-1

			onlyBody_img1 = img1[bound_U:bound_D, bound_L:bound_R,:]
			onlyBody_img2 = img2[bound_U:bound_D, bound_L:bound_R,:]
			onlyBody_img3 = img3[bound_U:bound_D, bound_L:bound_R,:]
			# print "search area", bound_U,bound_D,bound_L,bound_R

			u1, v1, u2, v2 = acc_deep_flow(onlyBody_img1, onlyBody_img2, onlyBody_img3)
			# print u1.shape

			drawImg_path = os.path.join(frame_path, obj_index, reNum(img2_num, True) + ".png")
			# print drawImg_path
			
			# flowImg_in = np.array(Image.open(drawImg_path))
			# flowImg_in = Image.fromarray(flowImg_in, mode='RGB')

			flowImg_in = np.zeros_like(img1)
			flowImg_in[..., 0] = mask2
			flowImg_in[..., 1] = mask2
			flowImg_in[..., 2] = mask2
			flowImg_in = Image.fromarray(flowImg_in, mode='RGB')
			
			heel = Image.open(drawImg_path)
			# heel = Image.fromarray(heel, mode='RGB')

			flow_count, flowImg_out, heel_img, heelRow, heelCol = heel_strike(mask2, 
				row, col, bound_U, bound_D, bound_L, bound_R, u1, v1, u2, v2, radAcc_thres, 
				flowImg_in, heel, 'L')
			
			save_name = os.path.join(save_path, "heel")
			checkDir(save_name)
			save_name = os.path.join(save_name, str(img2_num) + ".png")
			heel_img = Image.fromarray(heel_img)
			heel_img.save(save_name)
				
			# flowImg_out.save(save_name)
			# flowImg_out = np.array(flowImg_out)
			# plt.imsave(save_name, flowImg_out, cmap='jet')

				
		# ws['A'+str(img2_num)] = flow_count
		# ws['C'+str(img2_num)] = heelCol
		# ws['D'+str(img2_num)] = heelRow

		# wb.save(wbname)


if __name__ == "__main__":

	radAcc_thres = 3
	obj_index = "008a013s00L"

	frame_path = "/Volumes/2TB/dataset/gait/SOTON/people_frame"
	rmBG_path = "/Volumes/2TB/dataset/gait/SOTON/rmBG2"
	mask_path = "/Volumes/2TB/dataset/gait/SOTON/sil2"
	save_path = "/Users/user/Desktop"

	main(radAcc_thres, frame_path, rmBG_path, mask_path, save_path)