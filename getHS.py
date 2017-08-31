from openpyxl import Workbook
import numpy as np
from acc_deep_flow import acc_deep_flow
from heel_strike import heel_strike
from PIL import Image
import os
import os.path
from checkDir import checkDir
from heelStrikeDatabase import reNum


def maskCoords(mask):
	maskShape = mask.shape

	if len(maskShape) == 2:
		coords = np.argwhere(mask == 255)
		# print coords
	else:
		coords = np.argwhere(mask == [255,255,255])
		# print coords

	return coords


if __name__ == "__main__":
	threshold = 1

	rmBG_dataPath = "/Volumes/2TB/HSdata/SOTONrmBGnoised/gaussian"
	mask_dataPath = "/Volumes/2TB/HSdata/SOTON/silhouette"
	savePath = "/Volumes/2TB/HSres/acc_gaussian"
	wb = Workbook()
	wbname = "/Volumes/2TB/excel/acc/acc_gaussian.xlsx"

	for maskPath, dirnames, data in os.walk(mask_dataPath):
		folders = maskPath.split('/')

		if len(folders) == 7:
			frames = filter(lambda x: x.lower().endswith(('.png')), data)
			# ang = folders[-2]
			# obj = folders[-1]
			# print ang
			# print obj
			database, noiselevel, obj = folderName.split('_')
			print noiselevel, obj
			
			ws = wb.create_sheet(title=obj)
			
			framesinFolder = len(frames)
			framesLimtaion = 0
 			for img in frames:
 				framesLimtaion += 1
 				if framesLimtaion == framesinFolder-1:
 					break

 				print img
 				frame_num = os.path.splitext(img)[0]

 				parent = os.path.join(rmBG_dataPath, obj)
 				imgPath = os.path.join(parent, img)
				img1 = np.array(Image.open(imgPath))
				[row, col, dim] = img1.shape

				img2_num = int(frame_num)+1
				img2Path = os.path.join(parent, reNum(img2_num))+".png"
				
				while  os.path.isfile(img2Path) == False:
					img2_num += 1
					img2Path = os.path.join(parent, reNum(img2_num))+".png"
				img2 = np.array(Image.open(img2Path))


				img3_num = int(img2_num)+1
				img3Path = os.path.join(parent, reNum(img3_num))+".png"

				while  os.path.isfile(img3Path) == False:
					img3_num += 1
					img3Path = os.path.join(parent, reNum(img3_num))+".png"
				img3 = np.array(Image.open(img3Path))
					

				# img3_num = img2_num + 1
				# img3Path = os.path.join(parent, reNum(img3_num))+".png"

				# while os.path.isfile(img3Path):
				# 	img3 = np.array(Image.open(img2Path))
				# 	break       
				# else:
				# 	img3_num += 1
				# 	img3Path = os.path.join(parent, reNum(img3_num))+".png"

				
				# img3 = np.array(Image.open(img3Path))
				# while img3 == None:
				# 	img3_num += 1
				# 	img3Path = os.path.join(parent, reNum(img3_num))+".png"
				# 	img3 = np.array(Image.open(img3Path))


				
				mask1Path = os.path.join(maskPath, img)
				mask1 = np.array(Image.open(mask1Path))
				coords = maskCoords(mask1)
				# print coords.shape
				
				sil1R_min = min(coords[:,0])
				sil1R_max = max(coords[:,0])
				# print "sil1", sil1R_min, sil1R_max
				
				sil1C_min = min(coords[:,1])
				sil1C_max = max(coords[:,1])
				# print "sil1", sil1C_min, sil1C_max

				mask2Path = os.path.join(maskPath, reNum(img2_num))+".png"
				mask2 = np.array(Image.open(mask2Path))
				coords = maskCoords(mask2)
				# print coords

				sil2R_min = min(coords[:,0])
				sil2R_max = max(coords[:,0])
				# print "sil2", sil2R_min, sil2R_max

				sil2C_min= min(coords[:,1])
				sil2C_max = max(coords[:,1])
				# print "sil2", sil2C_min, sil2C_max
				
				mask3Path = os.path.join(maskPath, reNum(img3_num))+".png"
				mask3 = np.array(Image.open(mask3Path))
				coords = maskCoords(mask3)
				# print coords

				sil3R_min = min(coords[:,0])
				sil3R_max = max(coords[:,0])
				# print "sil3", sil3R_min, sil3R_max

				sil3C_min = min(coords[:,1])
				sil3C_max = max(coords[:,1])
				# print "sil3", sil3C_min, sil3C_max


				bound_L = min(sil1C_min, sil2C_min, sil3C_min) - 50
				if bound_L < 0:
					bound_L = 0
				
				bound_R = max(sil1C_max, sil2C_max, sil3C_max) + 50
				if bound_R >= col:
					bound_R = col-1


				bound_U = min(sil1R_min, sil2R_min, sil3R_min) - 50
				if bound_U < 0:
					bound_U = 0

				bound_D = max(sil1R_max, sil2R_max, sil3R_max) + 50
				if bound_D >= row:
					bound_D = row-1

				img1 = img1[bound_U:bound_D, bound_L:bound_R,:]
				img2 = img2[bound_U:bound_D, bound_L:bound_R,:]
				img3 = img3[bound_U:bound_D, bound_L:bound_R,:]
				# print "search area", bound_U,bound_D,bound_L,bound_R

				u1, v1, u2, v2 = acc_deep_flow(img1, img2, img3)
				# print u1.shape
				
				# framePath = os.path.join(frame_dataPath, ang, obj, reNum(img2_num))+".png"
				flow = np.array(Image.open(img2Path))
				flow = Image.fromarray(flow, mode='RGB')

				heel = np.array(Image.open(img2Path))
				heel = Image.fromarray(heel, mode='RGB')

				flow_count, flow_img, heel_img, heelRow, heelCol = heel_strike(mask2, row, col, bound_U, bound_D,
					bound_L, bound_R, u1, v1, u2, v2, threshold, flow, heel, 'L')
				

				saveName = os.path.join(savePath)
				checkDir(saveName)
				saveName = os.path.join(saveName, obj)
				checkDir(saveName)
				saveName = os.path.join(saveName, img)
				
				flow_img.save(saveName)

				
				ws['A'+str(img2_num)] = flow_count
				ws['D'+str(img2_num)] = heelRow
				ws['C'+str(img2_num)] = heelCol

				wb.save(wbname)